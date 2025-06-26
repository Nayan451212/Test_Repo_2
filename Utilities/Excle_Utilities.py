import openpyxl

class Excel_utilities:

    @staticmethod
    def count_data_Excel(file_path,sheet_name):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
        return sheet.max_row

    @staticmethod
    def read_Excel_data(file_path,sheet_name,row_num,column_num):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
        return sheet.cell(row=row_num,column=column_num).value

    @staticmethod
    def write_Excel_data(file_path,sheet_name,row_num,column_num,data):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
        sheet.cell(row=row_num,column=column_num).value = data
        workbook.save(file_path)
